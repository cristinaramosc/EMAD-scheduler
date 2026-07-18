"""Exporta i importa un únic full de càlcul (.xlsx) amb totes les dades
acadèmiques (professors, grups, assignatures i aules), com a alternativa
a treballar directament amb l'arxiu FET.

Unitat de temps: sempre en HORES, en format decimal (p.ex. 1,5 = una hora
i mitja). El mínim per sessió és 1 hora.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from backend.repositories.academic_data_repository import AcademicDataRepository
except ModuleNotFoundError:  # pragma: no cover
    from repositories.academic_data_repository import AcademicDataRepository


HEADER_FILL = PatternFill(start_color="263447", end_color="263447", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FONT = Font(italic=True, color="667085")

TEACHER_COLUMNS = [
    ("name", "Nom"),
    ("short_name", "Nom curt"),
    ("active", "Actiu (Sí/No)"),
    ("no_gaps", "Sense buits (Sí/No)"),
    ("max_hours_per_day", "Màx. hores/dia"),
    ("max_consecutive_hours", "Màx. hores consecutives"),
    ("preferred_availability", "Disponibilitat preferida (franges separades per comes, p.ex. Dilluns 8:00, Dilluns 8:30)"),
    ("unavailable_slots", "Franges no disponibles (franges separades per comes)"),
]

GROUP_COLUMNS = [
    ("name", "Nom del grup"),
    ("active", "Actiu (Sí/No)"),
    ("no_gaps", "Sense buits (Sí/No)"),
    ("max_hours_per_day", "Màx. hores/dia"),
    ("max_consecutive_hours", "Màx. hores consecutives"),
    ("preferred_availability", "Disponibilitat preferida (franges separades per comes)"),
    ("unavailable_slots", "Franges no disponibles (franges separades per comes)"),
]

SUBJECT_COLUMNS = [
    ("subject", "Nom de l'assignatura"),
    ("group", "Grup"),
    ("teacher", "Professor/s (separats per coma si són diversos)"),
    ("weekly_hours", "Durada total (hores, p.ex. 1,5 = 1h 30min)"),
    ("max_session_days", "Màxim de dies en què es pot repartir (mínim 1h per sessió)"),
    ("fixed_day", "Dia fix (opcional)"),
    ("fixed_start", "Hora fixa (opcional, p.ex. 9:00)"),
]

ROOM_COLUMNS = [
    ("name", "Nom de l'aula"),
    ("active", "Activa (Sí/No)"),
]


def _bool_to_text(value: Any) -> str:
    return "Sí" if value else "No"


def _text_to_bool(value: Any) -> bool:
    text = str(value or "").strip().lower()
    return text in ("sí", "si", "s", "yes", "true", "1", "x")


def _slots_to_text(slots: Any) -> str:
    if not slots:
        return ""
    return ", ".join(str(slot) for slot in slots)


def _text_to_slots(value: Any) -> List[str]:
    text = str(value or "").strip()
    if not text:
        return []
    return [part.strip() for part in text.replace("\n", ",").split(",") if part.strip()]


def _format_hours(value: Any) -> Any:
    if value in (None, ""):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    return int(number) if number == int(number) else number


def _write_sheet(workbook: Workbook, title: str, columns, rows: List[Dict[str, Any]], note: str = "") -> None:
    sheet = workbook.create_sheet(title=title)

    start_row = 1
    if note:
        sheet.cell(row=1, column=1, value=note).font = NOTE_FONT
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        start_row = 2

    for col_index, (_, label) in enumerate(columns, start=1):
        cell = sheet.cell(row=start_row, column=col_index, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row_offset, row in enumerate(rows, start=start_row + 1):
        for col_index, (key, _) in enumerate(columns, start=1):
            sheet.cell(row=row_offset, column=col_index, value=row.get(key, ""))

    for col_index in range(1, len(columns) + 1):
        sheet.column_dimensions[get_column_letter(col_index)].width = 26
    sheet.freeze_panes = sheet.cell(row=start_row + 1, column=1)


def _teacher_rows(repo: AcademicDataRepository) -> List[Dict[str, Any]]:
    restrictions = {item["teacher"]: item for item in repo.list_teacher_restrictions()}
    rows = []
    for teacher in repo.list_teachers():
        restriction = restrictions.get(teacher["name"], {})
        rows.append({
            "name": teacher.get("name", ""),
            "short_name": teacher.get("short_name", ""),
            "active": _bool_to_text(teacher.get("active", True)),
            "no_gaps": _bool_to_text(restriction.get("no_gaps")),
            "max_hours_per_day": restriction.get("max_hours_per_day", ""),
            "max_consecutive_hours": restriction.get("max_consecutive_hours", ""),
            "preferred_availability": _slots_to_text(restriction.get("preferred_availability")),
            "unavailable_slots": _slots_to_text(restriction.get("unavailable_slots")),
        })
    return rows


def _group_rows(repo: AcademicDataRepository) -> List[Dict[str, Any]]:
    restrictions = {item["group"]: item for item in repo.list_group_restrictions()}
    rows = []
    for group in repo.list_groups():
        restriction = restrictions.get(group["name"], {})
        rows.append({
            "name": group.get("name", ""),
            "active": _bool_to_text(group.get("active", True)),
            "no_gaps": _bool_to_text(restriction.get("no_gaps")),
            "max_hours_per_day": restriction.get("max_hours_per_day", ""),
            "max_consecutive_hours": restriction.get("max_consecutive_hours", ""),
            "preferred_availability": _slots_to_text(restriction.get("preferred_availability")),
            "unavailable_slots": _slots_to_text(restriction.get("unavailable_slots")),
        })
    return rows


def _subject_rows(repo: AcademicDataRepository) -> List[Dict[str, Any]]:
    rows = []
    for assignment in repo.active_canonical_assignments():
        rows.append({
            "subject": assignment.get("subject", ""),
            "group": assignment.get("group", ""),
            "teacher": assignment.get("teacher", ""),
            "weekly_hours": _format_hours(assignment.get("weekly_hours")),
            "max_session_days": assignment.get("max_session_days", ""),
            "fixed_day": assignment.get("fixed_day", ""),
            "fixed_start": assignment.get("fixed_start", ""),
        })
    return rows


def _room_rows(repo: AcademicDataRepository) -> List[Dict[str, Any]]:
    return [
        {"name": room.get("name", ""), "active": _bool_to_text(room.get("active", True))}
        for room in repo.list_rooms()
    ]


def build_workbook(repo: AcademicDataRepository | None, blank: bool) -> BytesIO:
    """Construeix el full de càlcul. Si blank=True, només hi ha les
    capçaleres (per començar de zero); si no, s'omple amb les dades actuals."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    teacher_rows = [] if blank else _teacher_rows(repo)
    group_rows = [] if blank else _group_rows(repo)
    subject_rows = [] if blank else _subject_rows(repo)
    room_rows = [] if blank else _room_rows(repo)

    _write_sheet(
        workbook, "Professors", TEACHER_COLUMNS, teacher_rows,
        note="Dades i restriccions horàries de cada professor. Format d'hora: Dilluns 8:00, Dilluns 8:30... (blocs de mitja hora).",
    )
    _write_sheet(
        workbook, "Grups", GROUP_COLUMNS, group_rows,
        note="Dades i restriccions horàries de cada grup d'alumnes.",
    )
    _write_sheet(
        workbook, "Assignatures", SUBJECT_COLUMNS, subject_rows,
        note="Durada sempre en HORES (decimal, p.ex. 1,5 = 1h30min). Mínim 1 hora per sessió.",
    )
    _write_sheet(
        workbook, "Aules", ROOM_COLUMNS, room_rows,
    )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _read_sheet_rows(workbook, title: str, columns) -> List[Dict[str, Any]]:
    if title not in workbook.sheetnames:
        return []
    sheet = workbook[title]

    header_row_index = None
    header_map: Dict[int, str] = {}
    for row in sheet.iter_rows(min_row=1, max_row=3):
        labels = [str(cell.value or "").strip() for cell in row]
        matches = sum(1 for _, label in columns if label in labels)
        if matches >= max(1, len(columns) // 2):
            header_row_index = row[0].row
            for cell in row:
                for key, label in columns:
                    if str(cell.value or "").strip() == label:
                        header_map[cell.column] = key
            break

    if header_row_index is None:
        return []

    rows = []
    for row in sheet.iter_rows(min_row=header_row_index + 1):
        if all(cell.value in (None, "") for cell in row):
            continue
        record: Dict[str, Any] = {}
        for cell in row:
            key = header_map.get(cell.column)
            if key:
                record[key] = cell.value
        if record:
            rows.append(record)
    return rows


def import_workbook(repo: AcademicDataRepository, file_bytes: bytes) -> Dict[str, int]:
    """Llegeix el full de càlcul unificat i substitueix les dades acadèmiques
    del repositori (professors, grups, assignatures, aules, restriccions)."""
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)

    teacher_rows = _read_sheet_rows(workbook, "Professors", TEACHER_COLUMNS)
    group_rows = _read_sheet_rows(workbook, "Grups", GROUP_COLUMNS)
    subject_rows = _read_sheet_rows(workbook, "Assignatures", SUBJECT_COLUMNS)
    room_rows = _read_sheet_rows(workbook, "Aules", ROOM_COLUMNS)

    for existing in list(repo.active_canonical_assignments()):
        repo.delete_canonical_assignment(existing["id"])
    for existing in list(repo.list_teachers()):
        repo.delete_teacher(existing["name"])
    for existing in list(repo.list_groups()):
        repo.delete_group(existing["name"])
    for existing in list(repo.list_rooms()):
        repo.delete_room(existing["name"])

    teachers_created = 0
    for row in teacher_rows:
        name = str(row.get("name") or "").strip()
        if not name:
            continue
        repo.create_teacher({"name": name, "short_name": row.get("short_name") or ""})
        teachers_created += 1
        if row.get("no_gaps") is not None or row.get("unavailable_slots") or row.get("preferred_availability"):
            repo.upsert_teacher_restriction({
                "teacher": name,
                "no_gaps": _text_to_bool(row.get("no_gaps")),
                "max_hours_per_day": row.get("max_hours_per_day") or None,
                "max_consecutive_hours": row.get("max_consecutive_hours") or None,
                "preferred_availability": _text_to_slots(row.get("preferred_availability")),
                "unavailable_slots": _text_to_slots(row.get("unavailable_slots")),
            })

    groups_created = 0
    for row in group_rows:
        name = str(row.get("name") or "").strip()
        if not name:
            continue
        repo.create_group({"name": name})
        groups_created += 1
        if row.get("no_gaps") is not None or row.get("unavailable_slots") or row.get("preferred_availability"):
            repo.upsert_group_restriction({
                "group": name,
                "no_gaps": _text_to_bool(row.get("no_gaps")),
                "max_hours_per_day": row.get("max_hours_per_day") or None,
                "max_consecutive_hours": row.get("max_consecutive_hours") or None,
                "preferred_availability": _text_to_slots(row.get("preferred_availability")),
                "unavailable_slots": _text_to_slots(row.get("unavailable_slots")),
            })

    rooms_created = 0
    for row in room_rows:
        name = str(row.get("name") or "").strip()
        if not name:
            continue
        repo.create_room({"name": name})
        rooms_created += 1

    assignments_created = 0
    for row in subject_rows:
        subject = str(row.get("subject") or "").strip()
        group = str(row.get("group") or "").strip()
        teacher = str(row.get("teacher") or "").strip()
        if not subject or not group:
            continue
        try:
            weekly_hours = float(str(row.get("weekly_hours") or "0").replace(",", "."))
        except ValueError:
            weekly_hours = 0.0
        repo.create_canonical_assignment({
            "teacher": teacher,
            "subject": subject,
            "group": group,
            "weekly_hours": weekly_hours,
            "max_session_days": row.get("max_session_days") or "",
            "fixed_day": str(row.get("fixed_day") or "").strip(),
            "fixed_start": str(row.get("fixed_start") or "").strip(),
        })
        assignments_created += 1

    return {
        "teachers": teachers_created,
        "groups": groups_created,
        "rooms": rooms_created,
        "assignments": assignments_created,
    }
